import openpyxl

book = openpyxl.load_workbook("table.xlsx")
sheetname = book.sheetnames[0]
for name in book.sheetnames:
    if 'TT' in name:
        sheetname = name
        break
sh = book[sheetname]

remove_rows = 0
for val in sh.values:
    if "Monday" in str(val[0]):
        break
    remove_rows += 1

sh.delete_rows(1, remove_rows)

CourseLength = dict()
for r in sh.merged_cells:
    CourseLength[(r.min_row, r.min_col)] = r.size["columns"]

Venue = ""
Day = 0
Section = ""
Periods = dict()  # Use a dictionary to store unique entries

for row, val in enumerate(sh.values):
    if val[1] == None:
        Day += 1
    else:
        Venue = val[1]
        for col in range(2, len(val)):
            period = val[col]
            if period != None:
                Section = period.split("(")[-1].split(")")[0]
                Course = period.split("(")[0].strip()
                stime = (col - 2) * 10
                if (row + 5, col + 1) in CourseLength:
                    etime = (((col - 2) + CourseLength[(row + 5, col + 1)]) * 10) + 10
                else:
                    etime = 0
                    continue

                # Use a unique key to check for duplicates
                key = (Course, Section, str(stime - 30), str(etime - 40), str(Day), Venue)
                Periods[key] = key

# Convert dictionary values to a list
unique_periods = list(Periods.values())

def compare(p):
    if "Lab" in p[0]:
        return p[0][:-4] + p[1]
    return p[0] + p[1]

unique_periods.sort(key=compare)

file = open("Data.js", "w")
file.write("Courses = [\n")
for P in unique_periods:
    file.write('\t["' + '","'.join(P) + '"],\n')
file.write("]\n")
file.close()
